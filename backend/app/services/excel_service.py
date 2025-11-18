"""
Excel service for import/export operations.

Handles Excel file generation and parsing for criteria, components, and results.
"""

import io
import logging
from typing import List, Dict, Any

try:
    import pandas as pd  # type: ignore
    PANDAS_AVAILABLE = True
except ImportError:
    pd = None  # type: ignore
    PANDAS_AVAILABLE = False

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def _workbook_to_bytes(workbook: Workbook) -> io.BytesIO:
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _auto_size_columns(ws: Worksheet):
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(value))
            except Exception:
                continue
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width


class ExcelService:
    """Service for Excel import/export operations."""
    
    @staticmethod
    def parse_criteria_excel(file_contents: bytes) -> List[Dict[str, Any]]:
        """
        Parse criteria from an Excel file.
        
        Args:
            file_contents: Raw bytes of Excel file
            
        Returns:
            List of criteria dictionaries
            
        Raises:
            ValueError: If required columns are missing
        """
        if not PANDAS_AVAILABLE:
            raise ValueError("Excel parsing requires pandas to be installed on the server.")

        df = pd.read_excel(io.BytesIO(file_contents))  # type: ignore
        
        required_columns = ['name', 'weight']
        if not all(col in df.columns for col in required_columns):
            raise ValueError(f"Excel file must contain columns: {', '.join(required_columns)}")
        
        criteria_list = []
        for _, row in df.iterrows():
            criterion_data = {
                'name': str(row['name']),
                'weight': float(row['weight']),
                'description': str(row.get('description', '')) if pd.notna(row.get('description')) else None,
                'unit': str(row.get('unit', '')) if pd.notna(row.get('unit')) else None,
                'higher_is_better': bool(row.get('higher_is_better', True)) if pd.notna(row.get('higher_is_better')) else True,
                'minimum_requirement': float(row.get('minimum_requirement')) if pd.notna(row.get('minimum_requirement')) else None,
                'maximum_requirement': float(row.get('maximum_requirement')) if pd.notna(row.get('maximum_requirement')) else None,
            }
            criteria_list.append(criterion_data)
        
        return criteria_list
    
    @staticmethod
    def export_criteria_excel(criteria: List[Any], project_name: str) -> io.BytesIO:
        """
        Export criteria to Excel file.
        
        Args:
            criteria: List of criterion model objects
            project_name: Name of project for filename
            
        Returns:
            BytesIO object with Excel file
        """
        columns = [
            'name',
            'description',
            'weight',
            'unit',
            'higher_is_better',
            'minimum_requirement',
            'maximum_requirement'
        ]
        data = [
            {
                'name': criterion.name,
                'description': criterion.description or '',
                'weight': criterion.weight,
                'unit': criterion.unit or '',
                'higher_is_better': criterion.higher_is_better,
                'minimum_requirement': criterion.minimum_requirement or '',
                'maximum_requirement': criterion.maximum_requirement or ''
            }
            for criterion in criteria
        ]
        
        if PANDAS_AVAILABLE and data:
            try:
                df = pd.DataFrame(data, columns=columns)  # type: ignore
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:  # type: ignore
                    df.to_excel(writer, index=False, sheet_name='Criteria')
                output.seek(0)
                return output
            except Exception as exc:
                logger.warning("Pandas criteria export failed, falling back to openpyxl: %s", exc)

        wb = Workbook()
        ws = wb.active
        ws.title = "Criteria"
        ws.append([col.replace("_", " ").title() for col in columns])
        for row in data:
            ws.append([row[col] for col in columns])
        _auto_size_columns(ws)
        return _workbook_to_bytes(wb)
    
    @staticmethod
    def parse_components_excel(file_contents: bytes) -> List[Dict[str, Any]]:
        """
        Parse components from an Excel file.
        
        Args:
            file_contents: Raw bytes of Excel file
            
        Returns:
            List of component dictionaries
            
        Raises:
            ValueError: If required columns are missing
        """
        if not PANDAS_AVAILABLE:
            raise ValueError("Excel parsing requires pandas to be installed on the server.")

        df = pd.read_excel(io.BytesIO(file_contents))  # type: ignore
        
        required_columns = ['manufacturer', 'part_number']
        if not all(col in df.columns for col in required_columns):
            raise ValueError(f"Excel file must contain columns: {', '.join(required_columns)}")
        
        components_list = []
        for _, row in df.iterrows():
            component_data = {
                'manufacturer': str(row['manufacturer']),
                'part_number': str(row['part_number']),
                'description': str(row.get('description', '')) if pd.notna(row.get('description')) else None,
                'datasheet_url': str(row.get('datasheet_url', '')) if pd.notna(row.get('datasheet_url')) else None,
                'availability': row.get('availability', 'in_stock') if pd.notna(row.get('availability')) else 'in_stock',
            }
            components_list.append(component_data)
        
        return components_list
    
    @staticmethod
    def export_components_excel(components: List[Any], project_name: str) -> io.BytesIO:
        """
        Export components to Excel file.
        
        Args:
            components: List of component model objects
            project_name: Name of project for filename
            
        Returns:
            BytesIO object with Excel file
        """
        columns = [
            'manufacturer',
            'part_number',
            'description',
            'datasheet_url',
            'availability',
            'source'
        ]
        data = [
            {
                'manufacturer': component.manufacturer,
                'part_number': component.part_number,
                'description': component.description or '',
                'datasheet_url': component.datasheet_url or '',
                'availability': getattr(component.availability, "value", component.availability),
                'source': getattr(component.source, "value", component.source),
            }
            for component in components
        ]
        
        if PANDAS_AVAILABLE and data:
            try:
                df = pd.DataFrame(data, columns=columns)  # type: ignore
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:  # type: ignore
                    df.to_excel(writer, index=False, sheet_name='Components')
                output.seek(0)
                return output
            except Exception as exc:
                logger.warning("Pandas component export failed, falling back to openpyxl: %s", exc)

        wb = Workbook()
        ws = wb.active
        ws.title = "Components"
        ws.append([col.replace("_", " ").title() for col in columns])
        for row in data:
            ws.append([row[col] for col in columns])
        _auto_size_columns(ws)
        return _workbook_to_bytes(wb)
    
    @staticmethod
    def export_full_trade_study(
        project: Any,
        components: List[Any],
        criteria: List[Any],
        results: List[Dict[str, Any]]
    ) -> io.BytesIO:
        """
        Export complete trade study to multi-sheet Excel file.
        
        Args:
            project: Project model object
            components: List of component model objects
            criteria: List of criterion model objects
            results: List of result dictionaries with scores
            
        Returns:
            BytesIO object with multi-sheet Excel file
        """
        if PANDAS_AVAILABLE:
            try:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:  # type: ignore
                    ExcelService._write_summary_sheet_pd(writer, project, components, criteria)
                    ExcelService._write_criteria_sheet_pd(writer, criteria)
                    ExcelService._write_components_sheet_pd(writer, components)
                    ExcelService._write_scores_sheet_pd(writer, components, criteria, results)
                    ExcelService._write_ranking_sheet_pd(writer, results)
                output.seek(0)
                return output
            except Exception as exc:
                logger.warning("Pandas trade study export failed, falling back to openpyxl: %s", exc)

        workbook = Workbook()
        ExcelService._write_summary_sheet_wb(workbook, project, components, criteria)
        ExcelService._write_criteria_sheet_wb(workbook, criteria)
        ExcelService._write_components_sheet_wb(workbook, components)
        ExcelService._write_scores_sheet_wb(workbook, components, criteria, results)
        ExcelService._write_ranking_sheet_wb(workbook, results)
        return _workbook_to_bytes(workbook)

    @staticmethod
    def _write_summary_sheet_pd(writer, project, components, criteria):
        summary_data = {
            'Field': ['Project Name', 'Component Type', 'Description', 'Status', 'Created Date', 'Total Components', 'Total Criteria'],
            'Value': [
                project.name,
                project.component_type,
                project.description or '',
                project.status.value if hasattr(project.status, "value") else project.status,
                project.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(project.created_at, 'strftime') else str(project.created_at),
                len(components),
                len(criteria)
            ]
        }
        df_summary = pd.DataFrame(summary_data)  # type: ignore
        df_summary.to_excel(writer, sheet_name='Summary', index=False)

    @staticmethod
    def _write_criteria_sheet_pd(writer, criteria):
        if not criteria:
            return
        criteria_data = [
            {
                'Name': c.name,
                'Description': c.description or '',
                'Weight': c.weight,
                'Unit': c.unit or '',
                'Higher is Better': c.higher_is_better,
                'Min Requirement': c.minimum_requirement or '',
                'Max Requirement': c.maximum_requirement or ''
            }
            for c in criteria
        ]
        df_criteria = pd.DataFrame(criteria_data)  # type: ignore
        df_criteria.to_excel(writer, sheet_name='Criteria', index=False)

    @staticmethod
    def _write_components_sheet_pd(writer, components):
        if not components:
            return
        components_data = [
            {
                'Manufacturer': c.manufacturer,
                'Part Number': c.part_number,
                'Description': c.description or '',
                'Datasheet URL': c.datasheet_url or '',
                'Availability': c.availability.value if hasattr(c.availability, "value") else c.availability,
                'Source': c.source.value if hasattr(c.source, "value") else c.source
            }
            for c in components
        ]
        df_components = pd.DataFrame(components_data)  # type: ignore
        df_components.to_excel(writer, sheet_name='Components', index=False)

    @staticmethod
    def _write_scores_sheet_pd(writer, components, criteria, results):
        if not (components and criteria and results):
            return
        scores_matrix = []
        for result in results:
            component = result["component"]
            row_data = {
                'Manufacturer': component.manufacturer,
                'Part Number': component.part_number
            }

            for criterion in criteria:
                score = result["score_dict"].get(criterion.id)
                row_data[f'{criterion.name} (Score)'] = score.score if score else 'N/A'
                row_data[f'{criterion.name} (Rationale)'] = score.rationale if score and score.rationale else ''
                if score and getattr(score, "raw_value", None) is not None:
                    row_data[f'{criterion.name} (Raw Value)'] = score.raw_value

            row_data['Total Weighted Score'] = result["total_score"]
            scores_matrix.append(row_data)

        if scores_matrix:
            df_scores = pd.DataFrame(scores_matrix)  # type: ignore
            df_scores.to_excel(writer, sheet_name='Detailed Scores', index=False)

    @staticmethod
    def _write_ranking_sheet_pd(writer, results):
        if not results:
            return
        ranking_data = [
            {
                'Rank': rank,
                'Manufacturer': result["component"].manufacturer,
                'Part Number': result["component"].part_number,
                'Total Weighted Score': result["total_score"],
                'Availability': result["component"].availability.value if hasattr(result["component"].availability, "value") else result["component"].availability
            }
            for rank, result in enumerate(results, 1)
        ]
        if ranking_data:
            df_ranking = pd.DataFrame(ranking_data)  # type: ignore
            df_ranking.to_excel(writer, sheet_name='Rankings', index=False)

    @staticmethod
    def _write_summary_sheet_wb(workbook: Workbook, project, components, criteria):
        ws = workbook.active
        ws.title = "Summary"
        ws.append(["Field", "Value"])
        rows = [
            ("Project Name", project.name),
            ("Component Type", project.component_type),
            ("Description", project.description or ""),
            ("Status", project.status.value if hasattr(project.status, "value") else project.status),
            ("Created Date", project.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(project.created_at, 'strftime') else str(project.created_at)),
            ("Total Components", len(components)),
            ("Total Criteria", len(criteria)),
        ]
        for row in rows:
            ws.append(row)
        _auto_size_columns(ws)

    @staticmethod
    def _write_criteria_sheet_wb(workbook: Workbook, criteria):
        ws = workbook.create_sheet("Criteria")
        ws.append(["Name", "Description", "Weight", "Unit", "Higher is Better", "Min Requirement", "Max Requirement"])
        for c in criteria:
            ws.append([
                c.name,
                c.description or "",
                c.weight,
                c.unit or "",
                c.higher_is_better,
                c.minimum_requirement or "",
                c.maximum_requirement or "",
            ])
        _auto_size_columns(ws)

    @staticmethod
    def _write_components_sheet_wb(workbook: Workbook, components):
        ws = workbook.create_sheet("Components")
        ws.append(["Manufacturer", "Part Number", "Description", "Datasheet URL", "Availability", "Source"])
        for c in components:
            ws.append([
                c.manufacturer,
                c.part_number,
                c.description or "",
                c.datasheet_url or "",
                c.availability.value if hasattr(c.availability, "value") else c.availability,
                c.source.value if hasattr(c.source, "value") else c.source,
            ])
        _auto_size_columns(ws)

    @staticmethod
    def _write_scores_sheet_wb(workbook: Workbook, components, criteria, results):
        if not (components and criteria and results):
            return
        ws = workbook.create_sheet("Detailed Scores")
        headers = ["Manufacturer", "Part Number"]
        for criterion in criteria:
            headers.extend([
                f"{criterion.name} (Score)",
                f"{criterion.name} (Rationale)",
                f"{criterion.name} (Raw Value)",
            ])
        headers.append("Total Weighted Score")
        ws.append(headers)

        for result in results:
            component = result["component"]
            row = [component.manufacturer, component.part_number]
            for criterion in criteria:
                score = result["score_dict"].get(criterion.id)
                row.append(score.score if score else 'N/A')
                row.append(score.rationale if score and score.rationale else '')
                row.append(score.raw_value if score and getattr(score, "raw_value", None) is not None else '')
            row.append(result["total_score"])
            ws.append(row)
        _auto_size_columns(ws)

    @staticmethod
    def _write_ranking_sheet_wb(workbook: Workbook, results):
        if not results:
            return
        ws = workbook.create_sheet("Rankings")
        ws.append(["Rank", "Manufacturer", "Part Number", "Total Weighted Score", "Availability"])
        for rank, result in enumerate(results, 1):
            ws.append([
                rank,
                result["component"].manufacturer,
                result["component"].part_number,
                result["total_score"],
                result["component"].availability.value if hasattr(result["component"].availability, "value") else result["component"].availability,
            ])
        _auto_size_columns(ws)


# Singleton instance
_excel_service = ExcelService()


def get_excel_service() -> ExcelService:
    """Get the Excel service singleton."""
    return _excel_service
